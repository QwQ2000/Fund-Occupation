import xlrd
import xlwt
from xlutils import copy
from datetime import date
from datetime import timedelta
import openpyxl

class Contract:
    def __init__(self,line,name,sales,rate,rec):
        self.sales,self.rec,self.name,self.rate,self.line = sales,rec,name,rate,line

    def getOccupation(self,endDate):
        currentSales = 0
        interest = 0
        if self.rec == []:
            return None,0,0
        occupation = self.rec[0][1]
        for i in range(1,len(self.rec)):
            if occupation >=0:
                interest += self.rate * occupation * (self.rec[i][0] - self.rec[i - 1][0]).days
            if self.rec[i][1] < 0:
                currentSales += -self.rec[i][1]
            if currentSales == self.sales:
                return self.rec[i][0],interest,0
            occupation += self.rec[i][1]
        if occupation > 0:
            interest += ((endDate - self.rec[i][0]).days + 1) * self.rate * occupation
        return None,interest,occupation

def readDate(workbook,sheet,i,j):
    return date(*xlrd.xldate_as_tuple(sheet.cell_value(i,j),workbook.datemode)[:3])

def readREValue(sheet,i1,j1,i2,j2):
    if not sheet.cell(i1,j1).ctype:
        return sheet.cell(i2,j2).value
    return -sheet.cell(i1,j1).value

wbName = '1.xlsx'
workbook = xlrd.open_workbook(wbName)
s1 = workbook.sheet_by_index(0)
s2 = workbook.sheet_by_index(1)
i = 1
contractDict = dict()
try:
    while 1:
        if not s1.cell(i,0).ctype:
            break
        name,sales,rate = s1.cell(i,0).value.encode('utf-8'),s1.cell(i,10).value,s1.cell(i,12).value
        contractDict[name] = Contract(i,name,sales,rate,[])
        i += 1
except IndexError:
    pass
i = 1
endDate = date.fromtimestamp(0)
try:
    while 1:
        if not s2.cell(i,0).ctype:
            break
        name,dt,val = s2.cell(i,0).value.encode('utf-8'),readDate(workbook,s2,i,1),readREValue(s2,i,3,i,4)
        endDate = max(endDate,dt)
        contractDict[name].rec.append([dt,val])
        i += 1
except IndexError:
    pass
workbook = copy.copy(workbook)
s1 = workbook.get_sheet(0)
for v in contractDict.values():
    v.rec.sort(key = lambda x:x[0])
    dt,interest,occupation = v.getOccupation(endDate)
    if dt != None:
        dateFormat = xlwt.XFStyle()
        dateFormat.num_format_str = 'yyyy/mm/dd'
        s1.write(v.line,11,dt,dateFormat)
    s1.write(v.line,13,occupation)
    s1.write(v.line,14,interest)
workbook.save('test.xlsx')